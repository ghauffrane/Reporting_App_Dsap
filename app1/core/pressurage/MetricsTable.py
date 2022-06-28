import openpyxl
from openpyxl import load_workbook
import pandas as pd
import os

class BasicMetrics: 
    def __init__(self, cycle_nbr, cycle_data, existing_variables, excel_dir):
        self.cycle_nbr = cycle_nbr
        self.cycle_data = cycle_data.copy()
        self.exist_vars = existing_variables

        self.time_col_name = "ts "
        self.vars_colnames = {"4300 Analog values - Inside membrane pressure": "Inside membrane pressure", 
                            "16014 Computation values - press net weight": "Press net weight", 
                            "16016 Computation values - sliding average of press net weight": "Sliding average of press net weight", 
                            "16026 Computation values - extracted must quantity": "Extracted must quantity", 
                            "16028 Computation values - loaded product quantity": "Loaded product quantity", 
                            "16034 Computation values - must extraction percentage": "Must extraction percentage"}

        self.output_path = os.path.join(excel_dir,"press_cycle{}_analysis.xlsx".format(self.cycle_nbr+1) )
    
    def write_data_excel(self):
        self.cycle_data.to_excel(self.output_path, sheet_name = "cycle_data", index = False)


    def get_cycle_duration(self):

        """ This method extracts the start and stop time of the cycle and the duration"""
        data_size = len(self.cycle_data)
        last_row_idx = data_size - 1
        start_time = self.cycle_data[self.time_col_name][0]
        stop_time = self.cycle_data[self.time_col_name][last_row_idx]
        duration_info = "{} - {}".format(start_time, stop_time)

        return duration_info
    
    def get_basic_metrics(self):

        """ Takes no input, it uses the vars__colnames
        Returns a dict of varnames and their corresponding metric values """

        metrics = dict()
        
        if "4300 Analog values - Inside membrane pressure" not in self.exist_vars:
            metrics["Inside membrane pressure"] = "NA"

        for vr in self.exist_vars: 
            if vr == "4300 Analog values - Inside membrane pressure": 
                key = self.vars_colnames[vr]
                metrics[key] = round(self.cycle_data[vr].mean(), 4)

            else: 
                key = self.vars_colnames[vr]
                metrics[key] = round(self.cycle_data[vr].max(), 4)

        return metrics 
    


    def make_table(self):

        cycle_duration = self.get_cycle_duration()

        metrics = self.get_basic_metrics()
        
        
        wb = load_workbook(self.output_path)
        # bold = wb.add_format({'bold': True, 'border': True, 'align': 'center'})
        # border = wb.add_format({'border': True, 'align': 'center'})
        
        table_sheet = wb.create_sheet("cycle_metrics")
        table_sheet.cell(1, 1, "Press Cycle order nubmer")
        table_sheet.cell(2, 1, "Press Cycle (Start - Stop)") 
        table_sheet.cell(3, 1, "Inside pressure membrane")
        table_sheet.cell(4, 1, "Press net weigh")
        table_sheet.cell(5, 1, "Sliding average of press net weight")
        table_sheet.cell(6, 1, "Extracted must quantity")
        table_sheet.cell(7, 1, "Must extraction perçentage")
        table_sheet.cell(8, 1, "Loaded product quantity")
        table_sheet.cell(1, 2, self.cycle_nbr+1)
        table_sheet.cell(2, 2, cycle_duration)

        if "4300 Analog values - Inside membrane pressure" not in self.exist_vars:
            table_sheet.cell(3, 2, "NA")
        else:
            table_sheet.cell(3, 2, metrics["Inside membrane pressure"])
            
        table_sheet.cell(4, 2, metrics["Press net weight"])
        table_sheet.cell(5, 2, metrics["Sliding average of press net weight"])
        table_sheet.cell(6, 2, metrics["Extracted must quantity"])
        table_sheet.cell(7, 2, metrics["Must extraction percentage"])
        table_sheet.cell(8, 2, metrics["Loaded product quantity"])

        # formatting
        table_sheet.column_dimensions["A"].width = 40
        table_sheet.column_dimensions["B"].width = 40
        
        wb.save(self.output_path)
        wb.close()

        