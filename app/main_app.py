# ========================================= IMPORTS ==================================================================== #
import csv
import sys
import os
import time
import numpy as np
from functools import partial
from PyQt5 import QtWidgets
from PyQt5 import QtCore
from PyQt5.QtGui import QPixmap, QMovie, QIcon
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, QMainWindow, QSplashScreen, QHeaderView
from PyQt5.uic import loadUi
from core import sis
from core.pressurage import load 
from core.pressurage import cycles_module
from core.pressurage import MetricsCharts
from core.pressurage import MetricsTable
from core.pressurage import press_schemas
from core import dataframeTable
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import datetime
# ========================================= END IMPORTS ==================================================================== #
class SplashScreen(QSplashScreen):
    def __init__(self):
        super(QSplashScreen, self).__init__()
        loadUi("./new_splash.ui", self)
        self.setWindowFlags(QtCore.Qt.WindowStaysOnTopHint| QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.move(QApplication.instance().desktop().screen().rect().center() - self.rect().center())


    def progress(self):
        for i in range(100):
            time.sleep(0.1)
            self.progressBar.setValue(i)
# ========================================= MAIN WINDOW APP START ========================================================== #
class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow,self).__init__()
        loadUi("./latest.ui",self)
        self.setWindowIcon(QIcon("./Assets/diagram.png"))
        self.centralwidget.setLayout(self.horizontalLayout)
        self.Cycles_list.clear()
        self.phases_list.clear()
        self.ListWidget.clear()
        self.chart_area.clear()
        
        self.setWindowTitle("Reporting App")
        self.statusBar().showMessage('Ready .. Set .. Go 🚀 🚀')
        self.statusBar().setStyleSheet("background-color : green")

        self.pressing_btn.clicked.connect(self.go_pressing_widget)
        self.filter_widget_btn.clicked.connect(self.go_filter_widget)
        self.expert_report_widget_btn.clicked.connect(self.go_expert_widget)
        self.sys_btn.clicked.connect(self.go_sys_widget)
        self.browse.clicked.connect(partial(self.browsefiles, "browse"))
        self.browse_3.clicked.connect(partial(self.browsefiles, "browse_3"))
        self.process_data.clicked.connect(self.processData)
        self.process_data_3.clicked.connect(self.process_sys_data)
        
        self.cleanup_btn.clicked.connect(self.CleanUp)



    # =========================== Stacked Widget Redirections Start ============================================================= #
    def go_pressing_widget(self): 
        self.stacked_widget.setCurrentIndex(0)

    def go_filter_widget(self):
        self.stacked_widget.setCurrentIndex(1)

    def go_expert_widget(self):
        self.stacked_widget.setCurrentIndex(2)
    
    def go_sys_widget(self): 
        self.stacked_widget.setCurrentIndex(3)
    # =========================== Stacked Widget Redirections End ============================================================= #

    # =========================== Press Functions Start ======================================================================= #
    
    def browsefiles(self, clickedBrowse):
        fname=QFileDialog.getOpenFileName(self, 'Open file', 'CSV (*.csv *.tsv)')  
        
        if clickedBrowse == "browse":
            self.textBrowser.setText(fname[0])
            self.file_path =  fname[0]
        elif clickedBrowse == "browse_3":
            self.textBrowser_3.setText(fname[0])
            self.sys_data_path = fname[0]
    
    def processData(self):

        self.data_process_prog_bar.setValue(0)
        self.this_session_data = press_schemas.DataSession()
        self.this_session_data.chosen_file_path = self.file_path   
        self.this_session_data.update_current_chart_path = None
        self.this_session_data.pdf_charts_filenames = []

        self.Cycles_list.clear()
        self.phases_list.clear()
        self.ListWidget.clear()
        self.chart_area.clear()
        self.data_process_prog_bar.setValue(5)
        csvloader = load.LoadData()
        self.data_process_prog_bar.setValue(20)
        
        try: 
            csvdata = csvloader.load_dt(self.this_session_data.chosen_file_path)
            self.this_session_data.dataframe = csvdata
            self.this_session_data.existing_vars = csvloader.check_variables()
            self.data_process_prog_bar.setValue(45)
            
        except Exception: 
            self.statusBar().showMessage('Issue with loading your csv file', 1000)
            self.statusBar().setStyleSheet("background-color : pink")
        
        # ============================================================================================================ #
        file_name = self.this_session_data.chosen_file_path.split("/")[-1].split(".")[0]
        self.this_session_data.dir_path_excel = os.path.join("./press_reports/excel_reports/", file_name)
        self.this_session_data.dir_path_pdf = os.path.join("./press_reports/pdf_reports/", file_name)
        
        if os.path.exists(self.this_session_data.dir_path_excel) == False:
            os.mkdir(self.this_session_data.dir_path_excel)

        if os.path.exists(self.this_session_data.dir_path_pdf) == False:
            os.mkdir(self.this_session_data.dir_path_pdf)
        self.data_process_prog_bar.setValue(55)
        # ============================================================================================================ #

        filling_status, dewater_status, press_status, dry_status = csvloader.get_status_columns()
        cycles_manager = cycles_module.Cycles(csvdata)
        self.this_session_data.ordered_cycles_indices, total_cycles = cycles_manager.separate_cycles(filling_status, dewater_status, press_status, dry_status )
        
        cycles_nbrs = [str(elt+1) for elt in range(total_cycles)]
        self.Cycles_list.addItems(cycles_nbrs)
        self.ListWidget.addItems(self.this_session_data.existing_vars)
        self.phases_list.addItems(["all", "filling", "dewatering", "pressing", "drying"])
        self.data_process_prog_bar.setValue(80)
        self.preview_chart.clicked.connect(self.previewChart)
        self.Cycles_list.view().pressed.connect(self.SummaryCycleTable)
        self.excelReportBTN.clicked.connect(self.cycle_excel_report)
        self.checkBox.clicked.connect(self.getCurrentChart)
        self.pdfReportBTN.clicked.connect(self.make_pdf_report)

        self.data_process_prog_bar.setValue(100)
        self.statusBar().showMessage("Data successfully processed. There are {} cycles in this data file.".format(total_cycles))
        self.statusBar().setStyleSheet("background-color : green")


    def previewChart(self): 
        # Reset checkbox each new chart preview
        self.checkBox.setChecked(False)
        current_cycle_nbr = int(self.Cycles_list.currentText()) - 1
        variables = [str(x.text()) for x in self.ListWidget.selectedItems()]
        phase = self.phases_list.currentText()
        img_out_dir = "./tmpImages"

        if len(variables)> 2: 
            self.statusBar().showMessage('You can select a maximum of 2 variables to display in chart.')
            self.statusBar().setStyleSheet("background-color : pink")
        
        if len(variables)==0:
            self.statusBar().showMessage('Please select a variable to make a chart preview.')
            self.statusBar().setStyleSheet("background-color : pink")
            
        else: 
            cycles_manager = cycles_module.Cycles(self.this_session_data.dataframe)
            cycle_data = cycles_manager.__getcycle__(current_cycle_nbr, self.this_session_data.ordered_cycles_indices)
            chartmaker = MetricsCharts.Charts(current_cycle_nbr, cycle_data, self.this_session_data.existing_vars)
            
            if len(variables)==2: 
                output_path = chartmaker.bichart(variables, phase, img_out_dir)
            elif len(variables)==1: 
                output_path = chartmaker.unichart(variables, phase, img_out_dir)
            
            self.this_session_data.update_current_chart_path = output_path

            pixmap = QPixmap(output_path)
            self.chart_area.setPixmap(pixmap)
            self.chart_area.setScaledContents(True)
            self.statusBar().showMessage(f"{variables} displayed.")
            self.statusBar().setStyleSheet("background-color : green")


    def SummaryCycleTable(self): 
        current_cycle_nbr = int(self.Cycles_list.currentText()) - 1
        cycles_manager = cycles_module.Cycles(self.this_session_data.dataframe)
        cycle_data = cycles_manager.__getcycle__(current_cycle_nbr, self.this_session_data.ordered_cycles_indices)
        TableHandler = MetricsTable.BasicMetrics(current_cycle_nbr, cycle_data, self.this_session_data.existing_vars, self.this_session_data.dir_path_excel)
        cycle_duration = TableHandler.get_cycle_duration()
        metrics = TableHandler.get_basic_metrics()

        self.lbl_cycle_nbr_value.setText(str(current_cycle_nbr+1))
        self.lbl_cycle_duration_value.setText(cycle_duration)
        self.lbl_membrane_pressure_value.setText(str(metrics["Inside membrane pressure"]))
        self.lbl_press_nweight_value.setText(str(metrics["Press net weight"]))
        self.lbl_sl_avg_press_nweight_value.setText(str(metrics["Sliding average of press net weight"]))
        self.label_2.setText(str(metrics["Extracted must quantity"]))
        self.label_4.setText(str(metrics["Must extraction percentage"])+" %")
        self.label_6.setText(str(metrics["Loaded product quantity"]))


    def cycle_excel_report(self):

        current_cycle_nbr = int(self.Cycles_list.currentText()) - 1
        cycles_manager = cycles_module.Cycles(self.this_session_data.dataframe)
        cycle_data = cycles_manager.__getcycle__(current_cycle_nbr, self.this_session_data.ordered_cycles_indices)
        TableHandler = MetricsTable.BasicMetrics(current_cycle_nbr, cycle_data, self.this_session_data.existing_vars, self.this_session_data.dir_path_excel)
        ChartHandler = MetricsCharts.Charts(current_cycle_nbr, cycle_data, self.this_session_data.existing_vars)
        
        TableHandler.write_data_excel()
        TableHandler.make_table()

        chart_out_dir = "./press_reports/tmpImages"

        for phase in ["all", "filling", "dewatering", "pressing", "drying"]: 
            wb = load_workbook(TableHandler.output_path)
            ws = wb.create_sheet(phase)
            row_pos = 1
            
            for vr in self.this_session_data.existing_vars: 
                img_pos = "A"+str(row_pos)
                img_path = ChartHandler.unichart([vr], phase, chart_out_dir)
                img = openpyxl.drawing.image.Image(img_path)
                # img.anchor(ws.cell(img_pos))
                ws.add_image(img, img_pos)
                wb.save(TableHandler.output_path)
                wb.close()
                row_pos += 54

            self.statusBar().showMessage(f"Charts for {phase} phase are successfully created..")
            self.statusBar().setStyleSheet("background-color : green")
        self.statusBar().showMessage(f"Excel report for cycle {current_cycle_nbr+1} is successfully created under press_reports/excel_reports/ sub directory." )

        # Clean temImages directory
        imgs = os.listdir(chart_out_dir)
        if len(imgs) > 0: 
            for im in imgs: 
                os.remove(os.path.join(chart_out_dir, im))

    
    def getCurrentChart(self):
        if self.checkBox.isChecked():
            self.this_session_data.pdf_charts_filenames.append(self.this_session_data.update_current_chart_path)


    def make_pdf_report(self):
        import PIL
        from PIL import Image
        Image.VERSION = PIL.__version__

        current_cycle_nbr = int(self.Cycles_list.currentText()) - 1
        cycles_manager = cycles_module.Cycles(self.this_session_data.dataframe)
        cycle_data = cycles_manager.__getcycle__(current_cycle_nbr, self.this_session_data.ordered_cycles_indices)
        TableHandler = MetricsTable.BasicMetrics(current_cycle_nbr, cycle_data, self.this_session_data.existing_vars, self.this_session_data.dir_path_excel)
        cycle_duration = TableHandler.get_cycle_duration()
        metrics = TableHandler.get_basic_metrics()
        data = [["Cycle number", current_cycle_nbr], ["Cycle duration", cycle_duration]]
        for key, val in metrics.items(): 
            data.append([key, val])
        DF = pd.DataFrame(data, columns = ["Metric name", "Metric value"])
        fig, ax =plt.subplots(figsize=(8,4))
        ax.axis('tight')
        ax.axis('off')
        ax.table(cellText=DF.values,colLabels=DF.columns,loc='center')
        plt.savefig(f"./press_reports/tmpImages/{current_cycle_nbr}_metricsTable.jpeg")
        
        pdf_path = f'{self.this_session_data.dir_path_pdf}/press_report_cycle{current_cycle_nbr+1}.pdf'
        cnv = canvas.Canvas(pdf_path, pagesize= A4)
        cnv.setTitle(f"press_report_cycle{current_cycle_nbr+1}")
        cnv.setFont('Helvetica-Bold', 16)
        cnv.drawImage("./Assets/logo_dellToff.jpeg", 10, 720, width = 100, height = 150)
        cnv.drawString(180, 800, "Machine: PRESSA CONTINUA 1 - 2021")
        cnv.drawString(180, 780, "S/N: 800191200650")
        cnv.drawString(180, 760, "Customer: Leeuwenkuil Vineyards")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cnv.drawString(30, 700, f"Report: {now}")
        cnv.drawImage(f"./press_reports/tmpImages/{current_cycle_nbr}_metricsTable.jpeg", 30, 350, width = 500, height = 350)
        cnv.setFont('Helvetica-Bold', 12)
        cnv.drawString(300, 10, f"{cnv.getPageNumber()}")
        cnv.showPage()

        if len(self.this_session_data.pdf_charts_filenames) > 0:
            for path in self.this_session_data.pdf_charts_filenames:
                
                cnv.drawImage(path, 10, 30, width = 600, height= 600)
                cnv.drawString(300, 10, f"{cnv.getPageNumber()}")
                cnv.showPage()

        cnv.save()
        self.statusBar().showMessage("PDF report for this cycle is successfully created.")
        self.statusBar().setStyleSheet("background-color : green")

    def CleanUp(self): 
        images_dir = "./tmpImages"
        images_paths = os.listdir(images_dir)
        if len(images_paths)>0: 
            for p in images_paths: 
                os.remove(os.path.join(images_dir, p))
        self.statusBar().showMessage("Cleanup is successfully finished.")


    def process_sys_data(self):
        self.data_process_prog_bar_3.setValue(0)
        self.this_sys_session_data = press_schemas.SysDataSession()
        self.this_sys_session_data.chosen_file_path = self.sys_data_path  
        self.data_process_prog_bar_3.setValue(5)
        csvloader2 = load.LoadData()
        self.data_process_prog_bar_3.setValue(20)
        
        try: 
            csvdata = csvloader2.load_dt(self.this_sys_session_data.chosen_file_path)
            
        except Exception: 
            self.statusBar().showMessage('Issue with loading your csv file', 1000)
            self.statusBar().setStyleSheet("background-color : pink")
        start = csvdata['ts '].iloc[0]
        end = csvdata['ts '].iloc[-1]
        self.label_9.setText(f"SIS Matrix: {start} - {end}")
        self.this_sys_session_data.sys_dataframe = csvdata.copy()
        try: 

            self.this_sys_session_data.sys_df = csvdata[self.this_sys_session_data.col_names].copy()
            DF = csvdata[self.this_sys_session_data.col_names].copy()

        except Exception as InvalidColumn:
            self.statusBar().showMessage(f'Issue with loading your csv file: {InvalidColumn}', 3000)
            self.statusBar().setStyleSheet("background-color : pink")

        self.this_sys_session_data.sys_df.columns = ['_id','ts','Pression','Phase3','DM','DDM','DFlux','RR']
        DF.columns = ['_id','ts','Pression','Phase3','DM','DDM','DFlux','RR']
        self.data_process_prog_bar_3.setValue(45)
        
        
        DF['NP']=0
        Compteur = 1
        for i in range(len(DF)-1): 
            if (DF["Phase3"][i] == 1) & (DF["Phase3"][i+1] == 0):
                DF["NP"][i] = Compteur
                Compteur = Compteur + 1
            else:
                DF["NP"][i]=Compteur
        DF["NP"][len(DF)-1]=DF["NP"][len(DF)-2]
        self.data_process_prog_bar_3.setValue(70)

        self.this_sys_session_data.sys_df =  DF.copy()
        filt = np.where((DF['Pression'] >= 0.2) & (DF['Phase3'] == 1))
        ddf = DF.loc[filt[0]]
        self.this_sys_session_data.sys_ddf = ddf.copy()
        
        Fin = ddf["NP"].max() - 1
        self.this_sys_session_data.fin = Fin
 
        self.data_process_prog_bar_3.setValue(85)
        
        sis_algo = sis.SIS()
        sys_df = sis_algo.get_sys_matrix(fin_df= self.this_sys_session_data.fin,
        df= self.this_sys_session_data.sys_df, 
        ddf= self.this_sys_session_data.sys_ddf)

        sys_model = dataframeTable.DataFrameModel(sys_df)
        self.matrixView.setModel(sys_model)
        self.matrixView.horizontalHeader().setStretchLastSection(True)
        self.matrixView.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch)

        self.data_process_prog_bar_3.setValue(100)

        self.statusBar().showMessage("Data successfully processed.")
        self.statusBar().setStyleSheet("background-color : green")

    # =========================== Press Functions End ======================================================================= #
    # ========================================= MAIN WINDOW APP START ========================================================= #



if __name__ == "__main__":
    app=QApplication(sys.argv)
    # splash = SplashScreen()
    # splash.show()
    # splash.progress()
    mainwindow=MainWindow()
    # splash.finish(mainwindow)
    mainwindow.show()
    sys.exit(app.exec_())